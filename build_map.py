"""
Read 강릉초중고위치좌표.xlsx and write map.html (Leaflet, embedded GeoJSON).
Run: python build_map.py
"""
from __future__ import annotations

import json
import math
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
XLSX_GLOB = "*.xlsx"
VWORLD_KEY_FILE = ROOT / "vworld_api_key.txt"
NAVER_TILE_VERSION = "1778232861"
NAVER_STYLE_JSON_URL = "https://map.pstatic.net/nrb/styles/basic.json?fmt=png"


def fetch_naver_tile_info() -> dict:
    fallback = {
        "tileVersion": NAVER_TILE_VERSION,
        "tileTemplate": (
            f"https://map.pstatic.net/nrb/styles/basic/{NAVER_TILE_VERSION}/"
            "{z}/{x}/{y}.png"
        ),
    }
    try:
        import urllib.request

        req = urllib.request.Request(
            NAVER_STYLE_JSON_URL,
            headers={
                "User-Agent": "Mozilla/5.0",
                "Referer": "https://map.naver.com/",
            },
        )
        with urllib.request.urlopen(req, timeout=8) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        version = str(data.get("version") or fallback["tileVersion"])
        tiles = data.get("tiles") or []
        template = tiles[0] if tiles else fallback["tileTemplate"]
        return {"tileVersion": version, "tileTemplate": template}
    except Exception:
        return fallback


def resolve_basemap() -> dict:
    """기본 배경: 네이버 일반지도. vworld/osm은 BASEMAP 분기에서만 사용."""
    naver = fetch_naver_tile_info()
    return {
        "provider": "naver",
        "key": "",
        "tileVersion": naver["tileVersion"],
        "tileTemplate": naver["tileTemplate"],
    }


def find_workbook() -> Path:
    paths = sorted(
        p for p in ROOT.glob(XLSX_GLOB) if not p.name.startswith("~$")
    )
    if not paths:
        raise FileNotFoundError(f"No .xlsx in {ROOT}")
    for p in paths:
        if "위치좌표" in p.name:
            return p
    return paths[0]


def row_to_feature(header: list[str], row: tuple) -> dict | None:
    d = {header[i]: row[i] for i in range(len(header)) if i < len(row)}
    lat_s, lon_s = d.get("위도"), d.get("경도")
    if lat_s is None or lon_s is None:
        return None
    try:
        lat = float(lat_s)
        lon = float(lon_s)
    except (TypeError, ValueError):
        return None
    if not (-90 <= lat <= 90) or not (-180 <= lon <= 180):
        return None
    if math.isnan(lat) or math.isnan(lon):
        return None
    props = {
        k: v
        for k, v in d.items()
        if k not in ("위도", "경도") and v is not None and v != ""
    }
    return {
        "type": "Feature",
        "geometry": {"type": "Point", "coordinates": [lon, lat]},
        "properties": props,
    }


def feature_dedupe_key(feature: dict) -> tuple:
    props = feature.get("properties") or {}
    school_id = props.get("학교ID")
    if school_id not in (None, ""):
        return ("id", str(school_id).strip())
    lon, lat = feature["geometry"]["coordinates"]
    name = str(props.get("학교명") or "").strip()
    return ("place", name, round(float(lat), 6), round(float(lon), 6))


def dedupe_features(features: list[dict]) -> list[dict]:
    seen: set[tuple] = set()
    deduped: list[dict] = []
    for feature in features:
        key = feature_dedupe_key(feature)
        if key in seen:
            continue
        seen.add(key)
        deduped.append(feature)
    return deduped


def build_geojson(path: Path) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    header = list(next(it))
    features: list[dict] = []
    for row in it:
        if not row or all(c is None for c in row):
            continue
        f = row_to_feature(header, row)
        if f:
            features.append(f)
    wb.close()
    return {"type": "FeatureCollection", "features": dedupe_features(features)}


# 남한(본토·제주·주요 도서)을 한 화면에 담기 위한 근사 범위.
SOUTH_KOREA_SW = (33.0, 124.7)
SOUTH_KOREA_NE = (38.7, 131.9)


def south_korea_map_bounds(_fc: dict | None = None) -> list[list[float]]:
    return [list(SOUTH_KOREA_SW), list(SOUTH_KOREA_NE)]


def initial_view_bounds(fc: dict | None = None) -> list[list[float]]:
    features = (fc or {}).get("features") or []
    if not features:
        return south_korea_map_bounds(fc)
    lons = [f["geometry"]["coordinates"][0] for f in features]
    lats = [f["geometry"]["coordinates"][1] for f in features]
    min_lat, max_lat = min(lats), max(lats)
    min_lon, max_lon = min(lons), max(lons)
    lat_span = max(0.01, max_lat - min_lat)
    lon_span = max(0.01, max_lon - min_lon)
    pad_lat = max(0.12, lat_span * 0.06)
    pad_lon = max(0.12, lon_span * 0.06)
    sw = [
        max(SOUTH_KOREA_SW[0], min_lat - pad_lat),
        max(SOUTH_KOREA_SW[1], min_lon - pad_lon),
    ]
    ne = [
        min(SOUTH_KOREA_NE[0], max_lat + pad_lat),
        min(SOUTH_KOREA_NE[1], max_lon + pad_lon),
    ]
    return [sw, ne]


def build_region_index(fc: dict) -> dict[str, list[str]]:
    index: dict[str, set[str]] = {}
    for feature in fc.get("features") or []:
        props = feature.get("properties") or {}
        sido = str(props.get("시도교육청명") or "").strip()
        office = str(props.get("교육지원청명") or "").strip()
        if not sido or not office:
            continue
        index.setdefault(sido, set()).add(office)
    return {sido: sorted(offices) for sido, offices in sorted(index.items())}


HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="ko">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>전국 학교 위치</title>
  <link rel="preconnect" href="https://fonts.googleapis.com" />
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin />
  <link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@400;500;600;700&display=swap" rel="stylesheet" />
  <link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"
    integrity="sha256-p4NxAoJBhIIN+hmNHrzRCf9tD/miZyoHS5obTRR9BMY=" crossorigin="" />
  <style>
    :root {
      --ink: #0f172a;
      --ink-soft: #334155;
      --ink-muted: #64748b;
      --surface: rgba(255, 255, 255, 0.78);
      --surface-strong: rgba(255, 255, 255, 0.94);
      --line: rgba(148, 163, 184, 0.28);
      --glow: rgba(99, 102, 241, 0.16);
      --shadow-xl: 0 28px 70px -32px rgba(15, 23, 42, 0.55);
      --shadow-lg: 0 18px 44px -20px rgba(15, 23, 42, 0.28);
      --shadow-md: 0 10px 28px -14px rgba(15, 23, 42, 0.18);
      --radius-xl: 24px;
      --radius-lg: 16px;
      --radius-md: 12px;
      --tier-kinder: #f97316;
      --tier-elm: #5b6eef;
      --tier-mid: #14b8a6;
      --tier-high: #e11d48;
      --tier-special: #8b5cf6;
    }
    * { box-sizing: border-box; }
    html {
      height: 100%;
      font-family: "Noto Sans KR", "Segoe UI", system-ui, sans-serif;
      -webkit-font-smoothing: antialiased;
      -moz-osx-font-smoothing: grayscale;
    }
    body {
      height: 100%; margin: 0;
      padding: clamp(12px, 2.5vw, 22px);
      font-family: inherit;
      color: var(--ink);
      background-color: #dbe4ee;
      background-image:
        radial-gradient(ellipse 90% 70% at 12% 0%, rgba(99, 102, 241, 0.22), transparent 58%),
        radial-gradient(ellipse 70% 55% at 100% 8%, rgba(20, 184, 166, 0.16), transparent 52%),
        radial-gradient(ellipse 60% 50% at 50% 100%, rgba(244, 63, 94, 0.08), transparent 55%),
        linear-gradient(165deg, #c9d4e3 0%, #dde5ef 38%, #edf1f7 72%, #f8fafc 100%);
      background-attachment: fixed;
    }
    .map-shell {
      position: fixed;
      inset: clamp(12px, 2.5vw, 22px);
      border-radius: var(--radius-xl);
      overflow: hidden;
      z-index: 0;
      background: #d7dee8;
      box-shadow:
        0 0 0 1px rgba(255, 255, 255, 0.55) inset,
        var(--shadow-xl),
        var(--shadow-md);
    }
    .map-shell::after {
      content: "";
      position: absolute;
      inset: 0;
      pointer-events: none;
      z-index: 650;
      border-radius: inherit;
      box-shadow:
        inset 0 0 0 1px rgba(255, 255, 255, 0.28),
        inset 0 -28px 48px -24px rgba(15, 23, 42, 0.12);
    }
    #map {
      height: 100%; width: 100%;
      background: linear-gradient(180deg, #e8edf4 0%, #d5dde8 100%);
    }
    .leaflet-container {
      font-family: inherit;
      background: #d5dde8;
    }
    .leaflet-tile-pane.tile-tune-map {
      filter: saturate(0.9) contrast(1.05) brightness(1.03) hue-rotate(-3deg);
    }
    .leaflet-top.leaflet-left {
      top: auto !important;
      bottom: clamp(16px, 3vw, 28px) !important;
      left: clamp(16px, 3vw, 28px) !important;
      margin: 0 !important;
    }
    .leaflet-bar {
      border: none !important;
      border-radius: var(--radius-md) !important;
      overflow: hidden;
      box-shadow: var(--shadow-md) !important;
    }
    .leaflet-bar a {
      width: 36px !important; height: 36px !important; line-height: 36px !important;
      font-size: 18px !important;
      color: var(--ink-soft) !important;
      border: none !important;
      background: var(--surface-strong) !important;
      transition: background 0.16s ease, color 0.16s ease;
    }
    .leaflet-bar a:hover { background: #fff !important; color: var(--ink) !important; }
    .leaflet-bar a.leaflet-disabled { color: #cbd5e1 !important; }
    .leaflet-control-zoom-in { border-bottom: 1px solid var(--line) !important; }
    .leaflet-control-attribution {
      background: var(--surface) !important;
      backdrop-filter: blur(14px) saturate(1.25);
      -webkit-backdrop-filter: blur(14px) saturate(1.25);
      color: var(--ink-muted) !important;
      font-size: 10px !important;
      padding: 5px 11px 6px !important;
      margin: 10px !important;
      border-radius: 999px !important;
      border: 1px solid rgba(255, 255, 255, 0.72) !important;
      box-shadow: var(--shadow-md) !important;
      max-width: calc(100% - 24px);
    }
    .leaflet-control-attribution a { color: #4f46e5 !important; font-weight: 600; }
    .leaflet-popup.school-popup-card .leaflet-popup-content-wrapper {
      border-radius: var(--radius-lg);
      background: var(--surface-strong);
      backdrop-filter: blur(16px) saturate(1.3);
      -webkit-backdrop-filter: blur(16px) saturate(1.3);
      box-shadow: var(--shadow-lg);
      border: 1px solid rgba(255, 255, 255, 0.88);
      overflow: hidden;
    }
    .leaflet-popup.school-popup-card .leaflet-popup-content {
      margin: 16px 18px;
      font-size: 13px;
      line-height: 1.65;
      color: var(--ink-soft);
    }
    .leaflet-popup.school-popup-card .leaflet-popup-content b {
      display: block;
      margin-bottom: 4px;
      font-size: 14px;
      color: var(--ink);
    }
    .leaflet-popup.school-popup-card .leaflet-popup-tip {
      box-shadow: none;
      background: var(--surface-strong);
    }
    .leaflet-popup-content-wrapper {
      border-radius: var(--radius-lg);
      box-shadow: var(--shadow-lg);
      border: 1px solid rgba(255, 255, 255, 0.88);
      overflow: hidden;
    }
    .leaflet-popup-content { margin: 14px 18px; font-size: 13px; line-height: 1.6; color: var(--ink-soft); }
    .leaflet-popup-tip { box-shadow: none; }
    .leaflet-popup-close-button {
      width: 32px !important; height: 32px !important;
      padding: 8px !important; font-size: 18px !important;
      color: #94a3b8 !important;
    }
    .leaflet-popup-close-button:hover { color: var(--ink-muted) !important; }
    .panel {
      position: absolute; z-index: 1000;
      top: clamp(22px, 4vw, 36px);
      left: clamp(58px, 8vw, 76px);
      padding: 18px 20px 16px;
      border-radius: var(--radius-lg);
      max-width: min(400px, calc(100vw - 32px));
      font-size: 13px;
      background: var(--surface);
      backdrop-filter: blur(18px) saturate(1.45);
      -webkit-backdrop-filter: blur(18px) saturate(1.45);
      border: 1px solid rgba(255, 255, 255, 0.82);
      box-shadow:
        0 0 0 1px rgba(148, 163, 184, 0.14),
        var(--shadow-lg);
    }
    .panel::before {
      content: "";
      position: absolute;
      top: 0; left: 18px; right: 18px; height: 3px;
      border-radius: 0 0 8px 8px;
      background: linear-gradient(90deg, #6366f1, #14b8a6, #f43f5e, #8b5cf6);
      opacity: 0.95;
    }
    .panel-kicker {
      margin: 2px 0 6px;
      font-size: 11px;
      font-weight: 700;
      letter-spacing: 0.08em;
      text-transform: uppercase;
      color: #6366f1;
    }
    .panel h1 {
      margin: 0 0 10px;
      font-size: 18px;
      font-weight: 700;
      letter-spacing: -0.04em;
      color: var(--ink);
      line-height: 1.2;
    }
    .panel p { margin: 0; color: var(--ink-muted); line-height: 1.6; font-size: 12px; font-weight: 400; }
    .region-filter {
      margin-top: 12px;
      display: grid;
      gap: 8px;
    }
    .region-field {
      display: grid;
      gap: 4px;
      font-size: 11px;
      font-weight: 600;
      color: var(--ink-soft);
    }
    .region-field input {
      width: 100%;
      border: 1px solid rgba(148, 163, 184, 0.35);
      border-radius: var(--radius-md);
      padding: 8px 10px;
      font-family: inherit;
      font-size: 12px;
      color: var(--ink);
      background: rgba(255, 255, 255, 0.9);
    }
    .region-field input:focus {
      outline: none;
      border-color: rgba(99, 102, 241, 0.55);
      box-shadow: 0 0 0 3px rgba(99, 102, 241, 0.14);
    }
    .btn-region {
      width: 100%;
      cursor: pointer;
      border: none;
      border-radius: var(--radius-md);
      padding: 9px 12px;
      font-family: inherit;
      font-size: 12px;
      font-weight: 700;
      color: #fff;
      background: linear-gradient(135deg, #4f46e5 0%, #6366f1 55%, #0d9488 100%);
      box-shadow: 0 6px 18px rgba(79, 70, 229, 0.24);
    }
    .btn-region:disabled {
      cursor: wait;
      opacity: 0.72;
    }
    .legend {
      margin-top: 14px;
      display: flex; flex-wrap: wrap;
      gap: 7px 8px;
    }
    .legend span.lg {
      display: inline-flex; align-items: center; gap: 7px;
      font-size: 11px; font-weight: 600;
      color: var(--ink-soft);
      padding: 6px 11px 6px 8px;
      border-radius: 999px;
      background: rgba(255, 255, 255, 0.66);
      border: 1px solid rgba(255, 255, 255, 0.9);
      box-shadow: 0 1px 2px rgba(15, 23, 42, 0.05);
    }
    .dot {
      width: 8px; height: 8px; border-radius: 50%;
      box-shadow: 0 0 0 2px rgba(255, 255, 255, 0.95), 0 1px 4px rgba(15, 23, 42, 0.18);
    }
    .leaflet-tooltip.school-label {
      border-radius: 999px;
      padding: var(--school-label-pad-v, 2px) var(--school-label-pad-h, 6px);
      font-size: var(--school-label-fs, 12px);
      font-weight: 700;
      letter-spacing: -0.03em;
      line-height: 1.25;
      box-shadow: 0 6px 18px rgba(15, 23, 42, 0.14), 0 1px 2px rgba(15, 23, 42, 0.08);
      pointer-events: none;
      backdrop-filter: blur(8px);
      -webkit-backdrop-filter: blur(8px);
    }
    .leaflet-tooltip.school-label.school-label-elm {
      background: rgba(238, 242, 255, 0.94);
      color: #1e3a8a;
      border: 1px solid rgba(91, 110, 239, 0.34);
    }
    .leaflet-tooltip.school-label.school-label-mid {
      background: rgba(236, 253, 245, 0.94);
      color: #115e59;
      border: 1px solid rgba(20, 184, 166, 0.34);
    }
    .leaflet-tooltip.school-label.school-label-high {
      background: rgba(255, 241, 242, 0.94);
      color: #9f1239;
      border: 1px solid rgba(225, 29, 72, 0.32);
    }
    .leaflet-tooltip.school-label.school-label-kinder {
      background: rgba(255, 247, 237, 0.94);
      color: #9a3412;
      border: 1px solid rgba(249, 115, 22, 0.34);
    }
    .leaflet-tooltip.school-label.school-label-special {
      background: rgba(245, 243, 255, 0.94);
      color: #5b21b6;
      border: 1px solid rgba(139, 92, 246, 0.34);
    }
    .leaflet-tooltip.school-label.school-label-other {
      background: rgba(248, 250, 252, 0.94);
      color: var(--ink-soft);
      border: 1px solid var(--line);
    }
    .leaflet-tooltip.school-label::before { display: none; }
    .school-label-lines {
      position: absolute;
      inset: 0;
      width: 100%;
      height: 100%;
      pointer-events: none;
      overflow: visible;
    }
    .school-label-line {
      stroke: rgba(51, 65, 85, 0.58);
      stroke-width: 1.25;
      fill: none;
      stroke-linecap: round;
    }
    .school-name-rail {
      position: absolute;
      top: clamp(8px, 1.5vw, 14px);
      right: clamp(8px, 1.5vw, 14px);
      bottom: clamp(52px, 10vh, 72px);
      width: max-content;
      max-width: calc(100% - 16px);
      min-width: 200px;
      max-height: calc(100% - 16px);
      z-index: 1001;
      pointer-events: auto;
      --rail-scale: 1;
      display: flex;
      flex-direction: column;
    }
    .school-name-rail--docked {
      bottom: auto !important;
      right: auto !important;
    }
    .school-name-rail--sized {
      bottom: auto !important;
      width: auto;
      max-width: calc(100% - 16px);
    }
    .school-name-rail--dragging,
    .school-name-rail--resizing {
      user-select: none;
    }
    .school-name-rail-drag {
      position: absolute;
      top: 0;
      left: 0;
      right: 0;
      z-index: 2;
      display: flex;
      align-items: center;
      justify-content: center;
      gap: 4px;
      height: calc(24px * var(--rail-scale, 1));
      margin: 0;
      padding: 0;
      border: none;
      border-radius: var(--radius-lg) var(--radius-lg) 0 0;
      background: rgba(255, 255, 255, 0.42);
      color: var(--ink-muted);
      font-family: inherit;
      font-size: calc(10px * var(--rail-scale, 1));
      font-weight: 700;
      letter-spacing: 0.04em;
      cursor: grab;
      touch-action: none;
    }
    .school-name-rail-drag::before {
      content: "";
      width: 28px;
      height: 4px;
      border-radius: 99px;
      background: linear-gradient(90deg, #cbd5e1 0%, #94a3b8 50%, #cbd5e1 100%);
    }
    .school-name-rail-drag:hover {
      color: var(--ink-soft);
      background: rgba(255, 255, 255, 0.58);
    }
    .school-name-rail--dragging .school-name-rail-drag,
    .school-name-rail-drag:active {
      cursor: grabbing;
    }
    .school-name-rail-inner {
      flex: 1 1 auto;
      min-height: 0;
      max-height: 100%;
      width: 100%;
      overflow: hidden;
      padding:
        calc(30px * var(--rail-scale, 1))
        calc(14px * var(--rail-scale, 1))
        calc(12px * var(--rail-scale, 1));
      border-radius: var(--radius-lg);
      background: var(--surface);
      backdrop-filter: blur(16px) saturate(1.35);
      -webkit-backdrop-filter: blur(16px) saturate(1.35);
      border: 1px solid rgba(255, 255, 255, 0.82);
      box-shadow:
        0 0 0 1px rgba(148, 163, 184, 0.12),
        var(--shadow-lg);
    }
    .school-name-rail-resize {
      position: absolute;
      right: 2px;
      bottom: 2px;
      z-index: 3;
      width: calc(18px * var(--rail-scale, 1));
      height: calc(18px * var(--rail-scale, 1));
      margin: 0;
      padding: 0;
      border: none;
      border-radius: 6px 0 var(--radius-lg) 0;
      background: rgba(255, 255, 255, 0.72);
      color: var(--ink-muted);
      cursor: nwse-resize;
      touch-action: none;
    }
    .school-name-rail-resize::before {
      content: "";
      position: absolute;
      right: 4px;
      bottom: 4px;
      width: 8px;
      height: 8px;
      border-right: 2px solid currentColor;
      border-bottom: 2px solid currentColor;
      opacity: 0.72;
    }
    .school-name-rail-resize:hover {
      color: var(--ink-soft);
      background: rgba(255, 255, 255, 0.92);
    }
    .school-name-rail--resizing .school-name-rail-resize,
    .school-name-rail-resize:active {
      cursor: nwse-resize;
    }
    .sn-section { margin-bottom: calc(14px * var(--rail-scale, 1)); }
    .sn-section:last-child { margin-bottom: 0; }
    .sn-h {
      margin: 0 0 6px;
      padding-left: 9px;
      border-left: 3px solid #94a3b8;
      font-size: calc(11px * var(--rail-scale, 1));
      font-weight: 700;
      letter-spacing: -0.02em;
      color: var(--ink-muted);
    }
    .sn-kinder .sn-h { border-left-color: #ea580c; color: #c2410c; }
    .sn-elm .sn-h { border-left-color: #4f6bed; color: #1d4ed8; }
    .sn-mid .sn-h { border-left-color: #0d9488; color: #0f766e; }
    .sn-high .sn-h { border-left-color: #c0264a; color: #9f1239; }
    .sn-special .sn-h { border-left-color: #7c3aed; color: #6d28d9; }
    .sn-flow {
      display: grid;
      grid-auto-flow: column;
      grid-auto-columns: minmax(calc(56px * var(--rail-scale, 1)), max-content);
      gap: calc(4px * var(--rail-scale, 1)) calc(6px * var(--rail-scale, 1));
      margin: 0;
      width: max-content;
    }
    .sn-kinder .sn-flow { grid-template-rows: repeat(1, auto); }
    .sn-elm .sn-flow { grid-template-rows: repeat(4, auto); }
    .sn-mid .sn-flow { grid-template-rows: repeat(2, auto); }
    .sn-high .sn-flow { grid-template-rows: repeat(2, auto); }
    .sn-special .sn-flow { grid-template-rows: repeat(1, auto); }
    .sn-chip {
      display: block;
      min-width: 0;
      font-size: calc(10px * var(--rail-scale, 1));
      font-weight: 700;
      line-height: 1.25;
      letter-spacing: -0.03em;
      color: var(--ink-soft);
      text-align: center;
      padding:
        calc(5px * var(--rail-scale, 1))
        calc(4px * var(--rail-scale, 1));
      border-radius: calc(8px * var(--rail-scale, 1));
      background: rgba(255, 255, 255, 0.72);
      border: 1px solid rgba(255, 255, 255, 0.92);
      box-shadow: 0 1px 2px rgba(15, 23, 42, 0.05);
      word-break: keep-all;
      overflow-wrap: anywhere;
      transition: transform 0.14s ease, box-shadow 0.14s ease, background 0.14s ease;
    }
    .sn-chip:hover {
      transform: translateY(-1px);
      background: rgba(255, 255, 255, 0.96);
      box-shadow: 0 4px 12px rgba(15, 23, 42, 0.08);
    }
    .sn-kinder .sn-chip { border-color: rgba(234, 88, 12, 0.28); }
    .sn-elm .sn-chip { border-color: rgba(79, 107, 237, 0.28); }
    .sn-mid .sn-chip { border-color: rgba(13, 148, 136, 0.28); }
    .sn-high .sn-chip { border-color: rgba(192, 38, 74, 0.28); }
    .sn-special .sn-chip { border-color: rgba(124, 58, 237, 0.28); }
    /* PNG 저장 직전에만 .map-shell에 붙음: 리스트 한 화면·촘촘히 */
    .map-shell--png-capture .school-name-rail {
      top: 6px;
      right: 6px;
      bottom: 6px;
      width: max-content;
      max-width: calc(100% - 12px);
      min-width: 100px;
      overflow: visible;
    }
    .map-shell--png-capture .school-name-rail-inner {
      max-height: none !important;
      overflow: visible;
      padding: 5px 6px 6px;
      border-radius: 10px;
    }
    .map-shell--png-capture .sn-flow {
      grid-auto-flow: column;
      grid-auto-columns: minmax(36px, max-content);
      gap: 2px 3px;
      width: max-content;
      max-width: none;
    }
    .map-shell--png-capture .sn-kinder .sn-flow { grid-template-rows: repeat(1, auto); }
    .map-shell--png-capture .sn-elm .sn-flow { grid-template-rows: repeat(4, auto); }
    .map-shell--png-capture .sn-mid .sn-flow { grid-template-rows: repeat(2, auto); }
    .map-shell--png-capture .sn-high .sn-flow { grid-template-rows: repeat(2, auto); }
    .map-shell--png-capture .sn-special .sn-flow { grid-template-rows: repeat(1, auto); }
    .map-shell--png-capture .sn-chip {
      font-size: 7px;
      font-weight: 700;
      padding: 2px 1px;
      line-height: 1.12;
      border-radius: 4px;
    }
    .map-shell--png-capture .sn-h {
      font-size: 7px;
      margin: 0 0 2px;
      padding-left: 5px;
      border-left-width: 2px;
    }
    .map-shell--png-capture .sn-section { margin-bottom: 4px; }
    .save-row { margin-top: 14px; }
    .btn-save {
      width: 100%;
      cursor: pointer;
      border: none;
      border-radius: var(--radius-md);
      padding: 11px 14px;
      font-family: inherit;
      font-size: 13px;
      font-weight: 700;
      letter-spacing: -0.02em;
      color: #fff;
      background: linear-gradient(135deg, #4f46e5 0%, #6366f1 42%, #0d9488 100%);
      box-shadow: 0 8px 22px rgba(79, 70, 229, 0.28);
      transition: transform 0.14s ease, box-shadow 0.14s ease, filter 0.14s ease;
    }
    .btn-save:hover:not(:disabled) {
      transform: translateY(-1px);
      filter: brightness(1.03);
      box-shadow: 0 12px 28px rgba(79, 70, 229, 0.34);
    }
    .btn-save:disabled {
      cursor: wait;
      opacity: 0.75;
    }
  </style>
</head>
<body>
  <div class="map-shell">
    <div id="map"></div>
    <aside class="school-name-rail" aria-label="학교명 목록">
      <button type="button" class="school-name-rail-drag" aria-label="학교명 목록 이동" title="드래그하여 이동">이동</button>
      <div class="school-name-rail-inner">
        <section class="sn-section sn-kinder">
          <h2 class="sn-h">유치원</h2>
          <div class="sn-flow" id="sn-list-kinder"></div>
        </section>
        <section class="sn-section sn-elm">
          <h2 class="sn-h">초등학교</h2>
          <div class="sn-flow" id="sn-list-elm"></div>
        </section>
        <section class="sn-section sn-mid">
          <h2 class="sn-h">중학교</h2>
          <div class="sn-flow" id="sn-list-mid"></div>
        </section>
        <section class="sn-section sn-high">
          <h2 class="sn-h">고등학교</h2>
          <div class="sn-flow" id="sn-list-high"></div>
        </section>
        <section class="sn-section sn-special">
          <h2 class="sn-h">특수학교</h2>
          <div class="sn-flow" id="sn-list-special"></div>
        </section>
      </div>
      <button type="button" class="school-name-rail-resize" aria-label="학교명 목록 크기 조절" title="드래그하여 크기 조절"></button>
    </aside>
  </div>
  <div class="panel">
    <p class="panel-kicker">전국 학교 지도</p>
    <h1>전국 학교 위치</h1>
    <p id="meta"></p>
    <div class="region-filter">
      <label class="region-field" for="filterSido">시도교육청
        <input id="filterSido" list="sidoOptions" type="search" placeholder="시도교육청 검색" autocomplete="off" />
        <datalist id="sidoOptions"></datalist>
      </label>
      <label class="region-field" for="filterOffice">지역교육청
        <input id="filterOffice" list="officeOptions" type="search" placeholder="교육지원청 검색" autocomplete="off" />
        <datalist id="officeOptions"></datalist>
      </label>
      <button type="button" class="btn-region" id="btnApplyRegion">지역 조회</button>
    </div>
    <div class="legend">
      <span class="lg"><i class="dot" style="background:#f97316"></i>유치원</span>
      <span class="lg"><i class="dot" style="background:#5b6eef"></i>초등</span>
      <span class="lg"><i class="dot" style="background:#14b8a6"></i>중등</span>
      <span class="lg"><i class="dot" style="background:#e11d48"></i>고등</span>
      <span class="lg"><i class="dot" style="background:#8b5cf6"></i>특수</span>
    </div>
    <div class="save-row">
      <button type="button" class="btn-save" id="btnSavePng" title="지도 영역을 고해상도 PNG로 저장합니다">
        이미지 저장 (고해상도)
      </button>
    </div>
  </div>
  <script src="https://cdnjs.cloudflare.com/ajax/libs/html2canvas/1.4.1/html2canvas.min.js" crossorigin="anonymous"></script>
  <script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"
    integrity="sha256-20nQCchB9co0qIjJZRGuk2/Z9VM+kNiyxNV1lvTlZBo=" crossorigin=""></script>
  <script>
  const GEO = __GEOJSON__;
  const MAP_BOUNDS = __MAP_BOUNDS__;
  const INITIAL_VIEW_BOUNDS = __INITIAL_VIEW_BOUNDS__;
  const REGION_INDEX = __REGION_INDEX__;
  const REGION_TOTAL = __REGION_TOTAL__;
  const BASEMAP = __BASEMAP__;
  function escapeHtml(s) {
    return String(s)
      .replace(/&/g, '&amp;')
      .replace(/</g, '&lt;')
      .replace(/>/g, '&gt;')
      .replace(/"/g, '&quot;');
  }
  function schoolTierFromGrade(grade) {
    const g = String(grade || '');
    if (g.indexOf('유치') !== -1) return 'kinder';
    if (g.indexOf('특수') !== -1) return 'special';
    if (g.indexOf('초') !== -1) return 'elm';
    if (g.indexOf('중') !== -1) return 'mid';
    if (g.indexOf('고') !== -1) return 'high';
    return 'other';
  }
  function schoolTierColor(tier) {
    if (tier === 'kinder') return '#f97316';
    if (tier === 'special') return '#8b5cf6';
    if (tier === 'elm') return '#5b6eef';
    if (tier === 'mid') return '#14b8a6';
    if (tier === 'high') return '#e11d48';
    return '#64748b';
  }
  function shortSchoolName(full) {
    if (!full) return '';
    let s = String(full);
    if (s.indexOf('초등학교') !== -1) s = s.replace('초등학교', '초');
    else if (s.indexOf('중학교') !== -1) s = s.replace('중학교', '중');
    else if (s.indexOf('고등학교') !== -1) s = s.replace('고등학교', '고');
    else if (s.indexOf('유치원') !== -1) s = s.replace('유치원', '유');
    else if (s.indexOf('특수학교') !== -1) s = s.replace('특수학교', '특');
    return s;
  }
  function fillSchoolNameRail(features) {
    const list = features || [];
    const RAIL_LIST_LIMIT = 400;
    if (!list.length) {
      ['sn-list-kinder', 'sn-list-elm', 'sn-list-mid', 'sn-list-high', 'sn-list-special']
        .forEach(function (id) {
          const el = document.getElementById(id);
          if (el) {
            el.innerHTML = '';
          }
        });
      return;
    }
    if (list.length > RAIL_LIST_LIMIT) {
      const note = '선택 지역 ' + list.length + '개 학교 · 목록은 확대 후 지도에서 확인';
      ['sn-list-kinder', 'sn-list-elm', 'sn-list-mid', 'sn-list-high', 'sn-list-special']
        .forEach(function (id) {
          const el = document.getElementById(id);
          if (!el) {
            return;
          }
          if (id === 'sn-list-elm') {
            el.innerHTML = '<span class="sn-chip" title="' + escapeHtml(note) + '">'
              + escapeHtml(note) + '</span>';
          } else {
            el.innerHTML = '';
          }
        });
      return;
    }
    const kinder = [];
    const elm = [];
    const mid = [];
    const high = [];
    const special = [];
    for (let i = 0; i < list.length; i++) {
      const p = list[i].properties || {};
      const tier = schoolTierFromGrade(p['학교급구분'] || '');
      const n = String(p['학교명'] || '').trim();
      if (!n) {
        continue;
      }
      if (tier === 'kinder') {
        kinder.push(n);
      } else if (tier === 'elm') {
        elm.push(n);
      } else if (tier === 'mid') {
        mid.push(n);
      } else if (tier === 'high') {
        high.push(n);
      } else if (tier === 'special') {
        special.push(n);
      }
    }
    const cmp = new Intl.Collator('ko').compare;
    function sortByShort(arr) {
      arr.sort(function (a, b) {
        return cmp(shortSchoolName(a), shortSchoolName(b));
      });
    }
    sortByShort(kinder);
    sortByShort(elm);
    sortByShort(mid);
    sortByShort(high);
    sortByShort(special);
    function fillFlow(id, arr) {
      const el = document.getElementById(id);
      if (!el) {
        return;
      }
      el.innerHTML = arr.map(function (fullName) {
        const sh = shortSchoolName(fullName);
        if (!sh) {
          return '';
        }
        return '<span class="sn-chip" title="' + escapeHtml(fullName) + '">'
          + escapeHtml(sh) + '</span>';
      }).join('');
    }
    fillFlow('sn-list-kinder', kinder);
    fillFlow('sn-list-elm', elm);
    fillFlow('sn-list-mid', mid);
    fillFlow('sn-list-high', high);
    fillFlow('sn-list-special', special);
  }
  function initSchoolRailPanel(shell) {
    var rail = document.querySelector('.school-name-rail');
    var handle = document.querySelector('.school-name-rail-drag');
    var resizeHandle = document.querySelector('.school-name-rail-resize');
    if (!rail || !handle || !resizeHandle || !shell) {
      return;
    }
    var MIN_RAIL_W = 200;
    var MIN_RAIL_H = 160;
    var BASE_RAIL_W = 300;
    var BASE_RAIL_H = 400;
    var dragging = false;
    var resizing = false;
    var offsetX = 0;
    var offsetY = 0;
    var resizeStartX = 0;
    var resizeStartY = 0;
    var resizeStartW = 0;
    var resizeStartH = 0;
    var inner = rail.querySelector('.school-name-rail-inner');
    var MIN_FIT_SCALE = 0.35;
    function railContentOverflows() {
      if (!inner) {
        return false;
      }
      return inner.scrollHeight > inner.clientHeight + 1
        || inner.scrollWidth > inner.clientWidth + 1;
    }
    function applyRailScaleToFit(preferredScale) {
      if (!inner) {
        return;
      }
      var lo = MIN_FIT_SCALE;
      var hi = preferredScale;
      var best = lo;
      for (var i = 0; i < 20; i++) {
        var mid = (lo + hi) / 2;
        rail.style.setProperty('--rail-scale', mid.toFixed(3));
        if (!railContentOverflows()) {
          best = mid;
          lo = mid;
        } else {
          hi = mid;
        }
      }
      rail.style.setProperty('--rail-scale', best.toFixed(3));
    }
    function updateRailScale(width, height) {
      var w = width || rail.offsetWidth;
      var h = height || rail.offsetHeight;
      var preferred = Math.min(w / BASE_RAIL_W, h / BASE_RAIL_H);
      preferred = Math.max(MIN_FIT_SCALE, Math.min(1.5, preferred));
      applyRailScaleToFit(preferred);
    }
    function clampRailSize(width, height) {
      var maxW = Math.max(MIN_RAIL_W, shell.clientWidth - 8);
      var maxH = Math.max(MIN_RAIL_H, shell.clientHeight - 8);
      return {
        width: Math.max(MIN_RAIL_W, Math.min(width, maxW)),
        height: Math.max(MIN_RAIL_H, Math.min(height, maxH))
      };
    }
    function clampRailPosition(left, top) {
      var maxLeft = Math.max(0, shell.clientWidth - rail.offsetWidth);
      var maxTop = Math.max(0, shell.clientHeight - rail.offsetHeight);
      return {
        left: Math.max(0, Math.min(left, maxLeft)),
        top: Math.max(0, Math.min(top, maxTop))
      };
    }
    function ensureDocked() {
      if (rail.classList.contains('school-name-rail--docked')) {
        return;
      }
      var shellRect = shell.getBoundingClientRect();
      var railRect = rail.getBoundingClientRect();
      applyRailPosition(railRect.left - shellRect.left, railRect.top - shellRect.top);
    }
    function applyRailPosition(left, top) {
      var pos = clampRailPosition(left, top);
      rail.style.left = pos.left + 'px';
      rail.style.top = pos.top + 'px';
      rail.style.right = 'auto';
      rail.style.bottom = 'auto';
      rail.classList.add('school-name-rail--docked');
    }
    function applyRailSize(width, height) {
      var size = clampRailSize(width, height);
      rail.style.width = size.width + 'px';
      rail.style.height = size.height + 'px';
      rail.classList.add('school-name-rail--sized');
      updateRailScale(size.width, size.height);
      if (rail.classList.contains('school-name-rail--docked')) {
        applyRailPosition(parseFloat(rail.style.left) || 0, parseFloat(rail.style.top) || 0);
      }
    }
    function onDragMove(e) {
      if (!dragging) {
        return;
      }
      var shellRect = shell.getBoundingClientRect();
      applyRailPosition(
        e.clientX - shellRect.left - offsetX,
        e.clientY - shellRect.top - offsetY
      );
      e.preventDefault();
    }
    function endDrag(e) {
      if (!dragging) {
        return;
      }
      dragging = false;
      rail.classList.remove('school-name-rail--dragging');
      if (handle.hasPointerCapture(e.pointerId)) {
        handle.releasePointerCapture(e.pointerId);
      }
      window.removeEventListener('pointermove', onDragMove);
      window.removeEventListener('pointerup', endDrag);
      window.removeEventListener('pointercancel', endDrag);
    }
    handle.addEventListener('pointerdown', function (e) {
      if (e.pointerType === 'mouse' && e.button !== 0) {
        return;
      }
      dragging = true;
      rail.classList.add('school-name-rail--dragging');
      ensureDocked();
      var shellRect = shell.getBoundingClientRect();
      var railRect = rail.getBoundingClientRect();
      offsetX = e.clientX - railRect.left;
      offsetY = e.clientY - railRect.top;
      handle.setPointerCapture(e.pointerId);
      window.addEventListener('pointermove', onDragMove);
      window.addEventListener('pointerup', endDrag);
      window.addEventListener('pointercancel', endDrag);
      e.preventDefault();
    });
    function onResizeMove(e) {
      if (!resizing) {
        return;
      }
      applyRailSize(
        resizeStartW + (e.clientX - resizeStartX),
        resizeStartH + (e.clientY - resizeStartY)
      );
      e.preventDefault();
    }
    function endResize(e) {
      if (!resizing) {
        return;
      }
      resizing = false;
      rail.classList.remove('school-name-rail--resizing');
      if (resizeHandle.hasPointerCapture(e.pointerId)) {
        resizeHandle.releasePointerCapture(e.pointerId);
      }
      window.removeEventListener('pointermove', onResizeMove);
      window.removeEventListener('pointerup', endResize);
      window.removeEventListener('pointercancel', endResize);
    }
    resizeHandle.addEventListener('pointerdown', function (e) {
      if (e.pointerType === 'mouse' && e.button !== 0) {
        return;
      }
      resizing = true;
      rail.classList.add('school-name-rail--resizing');
      ensureDocked();
      resizeStartX = e.clientX;
      resizeStartY = e.clientY;
      resizeStartW = rail.offsetWidth;
      resizeStartH = rail.offsetHeight;
      if (!rail.classList.contains('school-name-rail--sized')) {
        applyRailSize(resizeStartW, resizeStartH);
        resizeStartW = rail.offsetWidth;
        resizeStartH = rail.offsetHeight;
      }
      resizeHandle.setPointerCapture(e.pointerId);
      window.addEventListener('pointermove', onResizeMove);
      window.addEventListener('pointerup', endResize);
      window.addEventListener('pointercancel', endResize);
      e.preventDefault();
      e.stopPropagation();
    });
    window.addEventListener('resize', function () {
      if (rail.classList.contains('school-name-rail--sized')) {
        applyRailSize(rail.offsetWidth, rail.offsetHeight);
      } else {
        updateRailScale();
      }
      if (rail.classList.contains('school-name-rail--docked')) {
        applyRailPosition(parseFloat(rail.style.left) || 0, parseFloat(rail.style.top) || 0);
      }
    });
    window.requestAnimationFrame(function () {
      updateRailScale();
    });
  }
  (function () {
    var cachedGeoFeatures = null;
    var currentFeatureGroup = L.featureGroup();
    var schoolMarkers = [];
    initSchoolRailPanel(document.querySelector('.map-shell'));
    const maxBounds = L.latLngBounds(MAP_BOUNDS[0], MAP_BOUNDS[1]);
    const map = L.map('map', {
      maxBounds: maxBounds,
      maxBoundsViscosity: 1.0,
      preferCanvas: true,
      wheelPxPerZoomLevel: 150
    });
    var tilePane = map.getPanes().tilePane;
    if (tilePane && BASEMAP.provider !== 'naver') {
      tilePane.classList.add('tile-tune-map');
    }
    var tileUrl, tileOpts;
    if (BASEMAP.provider === 'naver') {
      var naverTemplate = BASEMAP.tileTemplate || (
        'https://map.pstatic.net/nrb/styles/basic/'
        + (BASEMAP.tileVersion || '1778232861') + '/{z}/{x}/{y}.png'
      );
      tileUrl = naverTemplate;
      tileOpts = {
        maxZoom: 21,
        minZoom: 0,
        bounds: maxBounds,
        attribution: '&copy; <a href="https://www.navercorp.com/" target="_blank" rel="noopener">NAVER</a>'
      };
    } else if (BASEMAP.provider === 'vworld' && BASEMAP.key) {
      tileUrl = 'https://api.vworld.kr/req/wmts/1.0.0/'
        + encodeURIComponent(BASEMAP.key) + '/Base/{z}/{y}/{x}.png';
      tileOpts = {
        maxZoom: 19,
        bounds: maxBounds,
        crossOrigin: true,
        attribution: '&copy; <a href="https://www.vworld.kr" target="_blank" rel="noopener">브이월드</a>(국토교통부)'
      };
    } else {
      tileUrl = 'https://tile.openstreetmap.org/{z}/{x}/{y}.png';
      tileOpts = {
        maxZoom: 19,
        bounds: maxBounds,
        crossOrigin: true,
        attribution: '&copy; <a href="https://www.openstreetmap.org/copyright">OpenStreetMap</a>'
      };
    }
    L.tileLayer(tileUrl, tileOpts).addTo(map);
    currentFeatureGroup.addTo(map);

    function loadAllFeatures() {
      if (cachedGeoFeatures) {
        return Promise.resolve(cachedGeoFeatures);
      }
      if (GEO && GEO.features) {
        cachedGeoFeatures = GEO.features;
        return Promise.resolve(cachedGeoFeatures);
      }
      if (window.location.protocol === 'file:') {
        return Promise.reject(new Error('embedded geo missing'));
      }
      return fetch('schools.geojson', { cache: 'no-store' })
        .then(function (resp) {
          if (!resp.ok) {
            throw new Error('geojson load failed');
          }
          return resp.json();
        })
        .then(function (fc) {
          cachedGeoFeatures = (fc && fc.features) ? fc.features : [];
          return cachedGeoFeatures;
        });
    }
    function resolveChoice(raw, options) {
      var value = String(raw || '').trim();
      if (!value) {
        return '';
      }
      if (options.indexOf(value) !== -1) {
        return value;
      }
      var matches = options.filter(function (item) {
        return item.indexOf(value) !== -1;
      });
      return matches.length === 1 ? matches[0] : '';
    }
    function setDatalistOptions(datalistId, options) {
      var datalist = document.getElementById(datalistId);
      if (!datalist) {
        return;
      }
      datalist.innerHTML = options.map(function (item) {
        return '<option value="' + escapeHtml(item) + '"></option>';
      }).join('');
    }
    function updateOfficeOptions() {
      var sidoInput = document.getElementById('filterSido');
      var sido = resolveChoice(sidoInput && sidoInput.value, Object.keys(REGION_INDEX));
      setDatalistOptions('officeOptions', REGION_INDEX[sido] || []);
    }
    function updateMeta(filteredCount, sido, office) {
      var baseNote = (BASEMAP.provider === 'naver')
        ? '배경: 네이버 일반지도'
        : (BASEMAP.provider === 'vworld' && BASEMAP.key)
        ? '배경: 브이월드(한글)'
        : '배경: OSM(지역명 한글 우선)';
      var scope = '';
      if (sido && office) {
        scope = sido + ' · ' + office;
      } else if (sido) {
        scope = sido;
      } else {
        scope = '전국 ' + REGION_TOTAL + '곳 중 지역 미선택';
      }
      document.getElementById('meta').textContent =
        scope + ' · 학교 ' + filteredCount + '곳 · ' + baseNote + ' · 학교: 엑셀';
    }
    function clearSchoolLayers() {
      currentFeatureGroup.clearLayers();
      schoolMarkers.length = 0;
      scheduleSchoolLabelLayout();
    }
    function renderSchoolFeatures(features) {
      clearSchoolLayers();
      if (!features.length) {
        fillSchoolNameRail([]);
        return;
      }
      var layers = [];
      for (var idx = 0; idx < features.length; idx++) {
        var f = features[idx];
        var coords = f.geometry.coordinates;
        var lon = coords[0];
        var lat = coords[1];
        var p = f.properties || {};
        var fullName = p['학교명'] || '';
        var label = shortSchoolName(fullName);
        var grade = String(p['학교급구분'] || '');
        var labelTier = schoolTierFromGrade(grade);
        var color = schoolTierColor(labelTier);
        var m = L.circleMarker([lat, lon], {
          radius: 6,
          weight: 2,
          color: '#ffffff',
          fillColor: color,
          fillOpacity: 0.94
        });
        var lines = [
          '<b>' + fullName + '</b>',
          p['학교급구분'] || '',
          (p['소재지도로명주소'] || p['소재지지번주소'] || '')
        ].filter(Boolean);
        m.bindPopup(lines.join('<br/>'), {
          className: 'school-popup-card',
          maxWidth: 280
        });
        if (label) {
          m.bindTooltip(label, {
            permanent: true,
            direction: 'center',
            offset: L.point(0, -12),
            className: 'school-label school-label-' + labelTier,
            opacity: 1,
            interactive: false
          });
          schoolMarkers.push(m);
        }
        currentFeatureGroup.addLayer(m);
        layers.push(m);
      }
      fillSchoolNameRail(features);
      var boundsGroup = L.featureGroup(layers);
      map.fitBounds(boundsGroup.getBounds().pad(0.08), schoolLabelFitBoundsOpts());
      scheduleSchoolLabelLayout();
    }
    function applyRegionFilter() {
      var btn = document.getElementById('btnApplyRegion');
      var sidoInput = document.getElementById('filterSido');
      var officeInput = document.getElementById('filterOffice');
      var sidoKeys = Object.keys(REGION_INDEX);
      var sido = resolveChoice(sidoInput && sidoInput.value, sidoKeys);
      var offices = REGION_INDEX[sido] || [];
      var office = resolveChoice(officeInput && officeInput.value, offices);
      if (!sido) {
        updateMeta(0, '', '');
        window.alert('시도교육청을 선택하거나 검색해 주세요.');
        return;
      }
      if (!office) {
        updateMeta(0, sido, '');
        window.alert('지역교육청(교육지원청)을 선택하거나 검색해 주세요.');
        return;
      }
      if (btn) {
        btn.disabled = true;
      }
      loadAllFeatures()
        .then(function (features) {
          var filtered = features.filter(function (feature) {
            var props = feature.properties || {};
            return String(props['시도교육청명'] || '').trim() === sido
              && String(props['교육지원청명'] || '').trim() === office;
          });
          renderSchoolFeatures(filtered);
          updateMeta(filtered.length, sido, office);
        })
        .catch(function () {
          window.alert(
            '학교 데이터를 불러오지 못했습니다.\\n'
            + 'build_map.py를 실행해 map.html을 다시 만든 뒤, map.html을 열어 주세요.'
          );
        })
        .finally(function () {
          if (btn) {
            btn.disabled = false;
          }
        });
    }
    function initRegionFilters() {
      setDatalistOptions('sidoOptions', Object.keys(REGION_INDEX));
      setDatalistOptions('officeOptions', []);
      var sidoInput = document.getElementById('filterSido');
      var officeInput = document.getElementById('filterOffice');
      var applyBtn = document.getElementById('btnApplyRegion');
      if (sidoInput) {
        sidoInput.addEventListener('input', updateOfficeOptions);
        sidoInput.addEventListener('change', updateOfficeOptions);
        sidoInput.addEventListener('keydown', function (e) {
          if (e.key === 'Enter') {
            applyRegionFilter();
          }
        });
      }
      if (officeInput) {
        officeInput.addEventListener('keydown', function (e) {
          if (e.key === 'Enter') {
            applyRegionFilter();
          }
        });
      }
      if (applyBtn) {
        applyBtn.addEventListener('click', applyRegionFilter);
      }
    }

    const ZOOM_LABELS_CENTER = 15;
    const ZOOM_LABELS_MIN = 8;
    const ZOOM_LABEL_FONT_MIN = 10;
    function schoolLabelFontSizePx(z) {
      if (z >= ZOOM_LABELS_CENTER) {
        return 12;
      }
      if (z <= ZOOM_LABEL_FONT_MIN) {
        return 8;
      }
      return 8 + (12 - 8) * (z - ZOOM_LABEL_FONT_MIN)
        / (ZOOM_LABELS_CENTER - ZOOM_LABEL_FONT_MIN);
    }
    function schoolLabelPaddingPx(z) {
      if (z >= ZOOM_LABELS_CENTER) {
        return { v: 3, h: 8 };
      }
      if (z <= ZOOM_LABEL_FONT_MIN) {
        return { v: 1, h: 4 };
      }
      var t = (z - ZOOM_LABEL_FONT_MIN) / (ZOOM_LABELS_CENTER - ZOOM_LABEL_FONT_MIN);
      return { v: 1 + 2 * t, h: 4 + 4 * t };
    }
    function applySchoolLabelZoomTypography(z) {
      var fs = schoolLabelFontSizePx(z);
      var pad = schoolLabelPaddingPx(z);
      var root = map.getContainer();
      root.style.setProperty('--school-label-fs', fs.toFixed(2) + 'px');
      root.style.setProperty('--school-label-pad-v', pad.v.toFixed(2) + 'px');
      root.style.setProperty('--school-label-pad-h', pad.h.toFixed(2) + 'px');
    }
    var schoolLabelLinksSvg = document.createElementNS('http://www.w3.org/2000/svg', 'svg');
    schoolLabelLinksSvg.setAttribute('class', 'school-label-lines');
    schoolLabelLinksSvg.setAttribute('aria-hidden', 'true');
    map.getPanes().tooltipPane.appendChild(schoolLabelLinksSvg);
    function resizeSchoolLabelLinks() {
      var size = map.getSize();
      schoolLabelLinksSvg.setAttribute('width', size.x);
      schoolLabelLinksSvg.setAttribute('height', size.y);
      schoolLabelLinksSvg.setAttribute('viewBox', '0 0 ' + size.x + ' ' + size.y);
    }
    const labelSpiral = [
      [0, -12], [12, -12], [-12, -12], [24, -12], [-24, -12],
      [0, 12], [12, 12], [-12, 12], [24, 12], [-24, 12],
      [36, -12], [-36, -12], [0, -28], [16, -28], [-16, -28],
      [36, 12], [-36, 12], [0, -44], [20, -44], [-20, -44],
      [48, -16], [-48, -16], [0, 32], [0, -56]
    ];
    function toMapRect(rect) {
      var mapRect = map.getContainer().getBoundingClientRect();
      return {
        left: rect.left - mapRect.left,
        top: rect.top - mapRect.top,
        right: rect.right - mapRect.left,
        bottom: rect.bottom - mapRect.top
      };
    }
    function anchorOnRect(rect, px, py) {
      var cx = (rect.left + rect.right) / 2;
      var cy = (rect.top + rect.bottom) / 2;
      var dx = cx - px;
      var dy = cy - py;
      if (dx === 0 && dy === 0) {
        return { x: cx, y: rect.bottom };
      }
      var hw = (rect.right - rect.left) / 2;
      var hh = (rect.bottom - rect.top) / 2;
      var adx = Math.abs(dx);
      var ady = Math.abs(dy);
      if (adx * hh > ady * hw) {
        var sx = dx > 0 ? 1 : -1;
        return { x: cx - sx * hw, y: cy - dy * (hw / adx) };
      }
      var sy = dy > 0 ? 1 : -1;
      return { x: cx - dx * (hh / ady), y: cy - sy * hh };
    }
    function clearSchoolLabelLines() {
      while (schoolLabelLinksSvg.firstChild) {
        schoolLabelLinksSvg.removeChild(schoolLabelLinksSvg.firstChild);
      }
    }
    function addSchoolLabelLine(x1, y1, x2, y2) {
      var line = document.createElementNS('http://www.w3.org/2000/svg', 'line');
      line.setAttribute('x1', x1.toFixed(2));
      line.setAttribute('y1', y1.toFixed(2));
      line.setAttribute('x2', x2.toFixed(2));
      line.setAttribute('y2', y2.toFixed(2));
      line.setAttribute('class', 'school-label-line');
      schoolLabelLinksSvg.appendChild(line);
    }
    function rectsOverlap(a, b, pad) {
      return !(
        a.right + pad < b.left ||
        a.left - pad > b.right ||
        a.bottom + pad < b.top ||
        a.top - pad > b.bottom
      );
    }
    function setSchoolLabelVisibility(visible) {
      schoolMarkers.forEach(function (mk) {
        var tip = mk.getTooltip();
        if (!tip) {
          return;
        }
        var el = tip.getElement();
        if (!el) {
          return;
        }
        el.style.display = visible ? '' : 'none';
        el.style.fontSize = '';
        el.style.padding = '';
        el.style.opacity = '';
        el.style.visibility = '';
      });
    }
    function layoutSchoolLabels() {
      if (!schoolMarkers.length) {
        clearSchoolLabelLines();
        return;
      }
      resizeSchoolLabelLinks();
      clearSchoolLabelLines();
      var z = map.getZoom();
      applySchoolLabelZoomTypography(z);
      if (z < ZOOM_LABELS_MIN) {
        setSchoolLabelVisibility(false);
        return;
      }
      setSchoolLabelVisibility(true);
      var viewBounds = map.getBounds().pad(0.35);
      var sorted = schoolMarkers.slice().sort(function (a, b) {
        var pa = map.latLngToContainerPoint(a.getLatLng());
        var pb = map.latLngToContainerPoint(b.getLatLng());
        if (pa.y !== pb.y) {
          return pa.y - pb.y;
        }
        return pa.x - pb.x;
      }).filter(function (mk) {
        return viewBounds.contains(mk.getLatLng());
      });
      var placed = [];
      var pad = 2;
      sorted.forEach(function (mk) {
        var tip = mk.getTooltip();
        if (!tip) {
          return;
        }
        var chosen = null;
        var chosenIndex = -1;
        var si;
        tip.options.direction = 'center';
        for (si = 0; si < labelSpiral.length; si++) {
          var off = labelSpiral[si];
          tip.options.offset = L.point(off[0], off[1]);
          tip.update();
          var el = tip.getElement();
          if (!el) {
            continue;
          }
          el.style.fontSize = '';
          el.style.padding = '';
          el.style.opacity = '';
          el.style.visibility = '';
          var r = toMapRect(el.getBoundingClientRect());
          var ok = true;
          var pi;
          for (pi = 0; pi < placed.length; pi++) {
            if (rectsOverlap(r, placed[pi], pad)) {
              ok = false;
              break;
            }
          }
          if (ok) {
            chosen = r;
            chosenIndex = si;
            break;
          }
        }
        if (!chosen) {
          var last = labelSpiral[labelSpiral.length - 1];
          tip.options.offset = L.point(last[0], last[1]);
          tip.update();
          chosenIndex = labelSpiral.length - 1;
          var fallbackEl = tip.getElement();
          if (fallbackEl) {
            fallbackEl.style.fontSize = '';
            fallbackEl.style.padding = '';
            fallbackEl.style.opacity = '';
            fallbackEl.style.visibility = '';
            chosen = toMapRect(fallbackEl.getBoundingClientRect());
          }
        }
        if (chosen) {
          placed.push(chosen);
        }
        if (chosenIndex > 0) {
          var markerPt = map.latLngToContainerPoint(mk.getLatLng());
          var anchor = anchorOnRect(chosen, markerPt.x, markerPt.y);
          addSchoolLabelLine(markerPt.x, markerPt.y, anchor.x, anchor.y);
        }
      });
    }
    var schoolLabelLayoutTimer = null;
    function scheduleSchoolLabelLayout() {
      if (schoolLabelLayoutTimer) {
        window.clearTimeout(schoolLabelLayoutTimer);
      }
      schoolLabelLayoutTimer = window.setTimeout(function () {
        schoolLabelLayoutTimer = null;
        layoutSchoolLabels();
        window.requestAnimationFrame(function () {
          layoutSchoolLabels();
        });
      }, 70);
    }
    map.on('zoomend', scheduleSchoolLabelLayout);
    map.on('moveend', scheduleSchoolLabelLayout);
    map.on('resize', function () {
      resizeSchoolLabelLinks();
      scheduleSchoolLabelLayout();
    });

    function schoolLabelFitBoundsOpts(extra) {
      var o = extra || {};
      return Object.assign({
        paddingTopLeft: L.point(8, 44),
        paddingBottomRight: L.point(8, 38)
      }, o);
    }
    const initialViewBounds = L.latLngBounds(
      INITIAL_VIEW_BOUNDS[0],
      INITIAL_VIEW_BOUNDS[1]
    );
    map.fitBounds(initialViewBounds, schoolLabelFitBoundsOpts());
    initRegionFilters();
    updateMeta(0, '', '');
    fillSchoolNameRail([]);

    var SAVE_SCALE = 3;
    function pad2(n) { return (n < 10 ? '0' : '') + n; }
    function makeFileName() {
      var d = new Date();
      return 'gangneung-school-map-'
        + d.getFullYear() + pad2(d.getMonth() + 1) + pad2(d.getDate())
        + '-' + pad2(d.getHours()) + pad2(d.getMinutes())
        + '.png';
    }
    function downloadPng(dataUrl) {
      var a = document.createElement('a');
      a.href = dataUrl;
      a.download = makeFileName();
      document.body.appendChild(a);
      a.click();
      document.body.removeChild(a);
    }
    function canvasToPng(canvas) {
      var url = canvas.toDataURL('image/png');
      if (!url || url.length < 32) {
        throw new Error('empty png');
      }
      return url;
    }
    function captureMapShell(opts) {
      var shell = document.querySelector('.map-shell');
      if (!shell || typeof html2canvas !== 'function') {
        return Promise.reject(new Error('no html2canvas'));
      }
      var hi = (opts.scale && opts.scale > 1) ? opts.scale : 1;
      var innerOpts = {};
      var k;
      for (k in opts) {
        if (Object.prototype.hasOwnProperty.call(opts, k)) {
          innerOpts[k] = opts[k];
        }
      }
      innerOpts.scale = 1;
      return html2canvas(shell, innerOpts).then(function (canvas) {
        if (hi <= 1) {
          return canvasToPng(canvas);
        }
        var out = document.createElement('canvas');
        out.width = Math.round(canvas.width * hi);
        out.height = Math.round(canvas.height * hi);
        var ctx = out.getContext('2d');
        ctx.imageSmoothingEnabled = true;
        ctx.imageSmoothingQuality = 'high';
        ctx.drawImage(
          canvas,
          0, 0, canvas.width, canvas.height,
          0, 0, out.width, out.height
        );
        return canvasToPng(out);
      });
    }
    function applyRailZoomForPngCapture() {
      var inner = document.querySelector('.school-name-rail-inner');
      var wrap = document.querySelector('.school-name-rail');
      if (!inner || !wrap) {
        return function () {};
      }
      inner.style.transform = '';
      inner.style.transformOrigin = '';
      var availH = Math.max(48, wrap.clientHeight);
      var availW = Math.max(80, wrap.clientWidth);
      var needH = inner.scrollHeight;
      var needW = inner.scrollWidth;
      var zr = 1;
      if (needH > availH) {
        zr = Math.min(zr, (availH / needH) * 0.98);
      }
      if (needW > availW) {
        zr = Math.min(zr, (availW / needW) * 0.98);
      }
      zr = Math.max(0.4, Math.min(1, zr));
      if (zr < 1) {
        inner.style.transformOrigin = 'top right';
        inner.style.transform = 'scale(' + zr + ')';
      }
      return function () {
        inner.style.transform = '';
        inner.style.transformOrigin = '';
      };
    }
    document.getElementById('btnSavePng').addEventListener('click', function () {
      var btn = this;
      if (btn.disabled) {
        return;
      }
      btn.disabled = true;
      var prevText = btn.textContent;
      btn.textContent = '저장 중…';
      var shell = document.querySelector('.map-shell');
      var rbSW = map.getBounds().getSouthWest();
      var rbNE = map.getBounds().getNorthEast();
      function restoreMapViewAfterPng() {
        map.fitBounds(L.latLngBounds(rbSW, rbNE), { animate: false });
        if (!maxBounds.contains(map.getBounds())) {
          map.fitBounds(maxBounds, schoolLabelFitBoundsOpts({ animate: false }));
        }
        map.invalidateSize(false);
        scheduleSchoolLabelLayout();
      }
      var base = {
        scale: SAVE_SCALE,
        logging: false,
        useCORS: true,
        allowTaint: false,
        backgroundColor: null,
        imageTimeout: 25000,
        scrollX: 0,
        scrollY: 0
      };
      window.setTimeout(function () {
        if (currentFeatureGroup.getLayers().length) {
          map.fitBounds(
            currentFeatureGroup.getBounds().pad(0.1),
            schoolLabelFitBoundsOpts({ animate: false })
          );
        }
        if (!maxBounds.contains(map.getBounds())) {
          map.fitBounds(maxBounds, schoolLabelFitBoundsOpts({ animate: false }));
        }
        map.invalidateSize(false);
        window.setTimeout(function () {
          if (shell) {
            shell.classList.add('map-shell--png-capture');
          }
          var restoreRailZoom = applyRailZoomForPngCapture();
          map.invalidateSize(false);
          layoutSchoolLabels();
          window.requestAnimationFrame(function () {
            layoutSchoolLabels();
            window.requestAnimationFrame(function () {
              layoutSchoolLabels();
              captureMapShell(base)
            .catch(function () {
              return captureMapShell({
                scale: SAVE_SCALE,
                logging: false,
                useCORS: false,
                allowTaint: false,
                backgroundColor: '#c7d2e0',
                imageTimeout: 25000,
                scrollX: 0,
                scrollY: 0,
                ignoreElements: function (el) {
                  return el.classList && el.classList.contains('leaflet-tile-pane');
                }
              });
            })
            .then(function (dataUrl) {
              downloadPng(dataUrl);
            })
            .catch(function () {
              window.alert(
                '이미지를 저장하지 못했습니다.\\n'
                + '브라우저에서 화면 캡처를 사용하거나, 페이지를 인쇄(PDF)해 보세요.'
              );
            })
            .finally(function () {
              if (shell) {
                shell.classList.remove('map-shell--png-capture');
              }
              restoreRailZoom();
              restoreMapViewAfterPng();
              btn.disabled = false;
              btn.textContent = prevText;
            });
            });
          });
        }, 650);
      }, 120);
    });
  })();
  </script>
</body>
</html>
"""


def main() -> None:
    wb_path = find_workbook()
    fc = build_geojson(wb_path)
    out_geo = ROOT / "schools.geojson"
    out_geo.write_text(json.dumps(fc, ensure_ascii=False, indent=2), encoding="utf-8")

    geo_js = json.dumps(fc, ensure_ascii=False)
    region_index = build_region_index(fc)
    region_index_js = json.dumps(region_index, ensure_ascii=False)
    region_total_js = str(len(fc["features"]))
    bounds = south_korea_map_bounds(fc)
    bounds_js = json.dumps(bounds)
    initial_bounds = initial_view_bounds(fc)
    initial_bounds_js = json.dumps(initial_bounds)
    basemap = resolve_basemap()
    basemap_js = json.dumps(basemap, ensure_ascii=False)
    html = (
        HTML_TEMPLATE.replace("__GEOJSON__", geo_js)
        .replace("__MAP_BOUNDS__", bounds_js)
        .replace("__INITIAL_VIEW_BOUNDS__", initial_bounds_js)
        .replace("__REGION_INDEX__", region_index_js)
        .replace("__REGION_TOTAL__", region_total_js)
        .replace("__BASEMAP__", basemap_js)
    )
    (ROOT / "map.html").write_text(html, encoding="utf-8")

    print(f"Workbook: {wb_path.name}")
    print("Basemap: Naver Street (일반지도)")
    print(f"Wrote {out_geo.name} ({len(fc['features'])} features)")
    print("Wrote map.html - open in a browser (double-click OK)")


if __name__ == "__main__":
    main()
